from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsAdmin
from courses.models import Dars, Fan, Daraja, Mavzu
from exams.audit import log_amal
from exams.models import Mashq, Savol, Javob, ListeningSavol, WritingTopshiriq, SpeakingTopshiriq

HEADER_FILL = PatternFill('solid', fgColor='173F35')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _style(ws):
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col in ws.columns:
        width = min(55, max(12, max(len(str(c.value or '')) for c in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = width


def _find_course(fan_name, level_name, topic_name, lesson_name):
    fan = Fan.objects.filter(nomi__iexact=str(fan_name).strip()).first()
    if not fan:
        raise ValueError(f"Fan topilmadi: {fan_name}")
    level = Daraja.objects.filter(fan=fan, nomi__iexact=str(level_name).strip()).first()
    if not level:
        raise ValueError(f"Daraja topilmadi: {fan_name} / {level_name}")
    topic, _ = Mavzu.objects.get_or_create(daraja=level, nomi=str(topic_name).strip(), defaults={'tartib': 0})
    lesson, _ = Dars.objects.get_or_create(mavzu=topic, sarlavha=str(lesson_name).strip(), defaults={'tartib': 0})
    return fan, level, topic, lesson


class ContentExcelExportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet('Mavzular')
        ws.append(['Fan', 'Daraja', 'Mavzu', 'Dars', 'Tushuntirish', 'Misollar', 'Tartib'])
        for lesson in Dars.objects.select_related('mavzu__daraja__fan').all():
            ws.append([lesson.mavzu.daraja.fan.nomi, lesson.mavzu.daraja.nomi, lesson.mavzu.nomi, lesson.sarlavha, lesson.tushuntirish_matn, lesson.misollar, lesson.tartib])
        _style(ws)

        ws = wb.create_sheet('TestSavollar')
        ws.append(['Fan', 'Daraja', 'Mavzu', 'Dars', 'Savol', 'A', 'B', 'C', 'D', 'Togri', 'Tartib'])
        for question in Savol.objects.select_related('mashq__dars__mavzu__daraja__fan').prefetch_related('javoblar'):
            lesson = question.mashq.dars
            answers = list(question.javoblar.all())[:4]
            values = [a.matn for a in answers] + [''] * (4 - len(answers))
            correct = next((chr(65 + i) for i, a in enumerate(answers) if a.togri), '')
            ws.append([lesson.mavzu.daraja.fan.nomi, lesson.mavzu.daraja.nomi, lesson.mavzu.nomi, lesson.sarlavha, question.matn, *values, correct, question.tartib])
        _style(ws)

        ws = wb.create_sheet('Listening')
        ws.append(['Fan', 'Daraja', 'Mavzu', 'Dars', 'AudioMatn', 'Savol', 'Variant1', 'Variant2', 'Variant3', 'Variant4', 'TogriJavob', 'Til'])
        for item in ListeningSavol.objects.select_related('dars__mavzu__daraja__fan'):
            lesson = item.dars
            variants = list(item.variantlar or [])[:4] + [''] * (4 - len(item.variantlar or []))
            ws.append([lesson.mavzu.daraja.fan.nomi, lesson.mavzu.daraja.nomi, lesson.mavzu.nomi, lesson.sarlavha, item.audio_matn, item.savol, *variants, item.togri_javob, item.til_kodi])
        _style(ws)

        ws = wb.create_sheet('Writing')
        ws.append(['Fan', 'Daraja', 'Mavzu', 'Dars', 'Topshiriq', 'MinimalSoz'])
        for item in WritingTopshiriq.objects.select_related('dars__mavzu__daraja__fan'):
            lesson = item.dars
            ws.append([lesson.mavzu.daraja.fan.nomi, lesson.mavzu.daraja.nomi, lesson.mavzu.nomi, lesson.sarlavha, item.matn, item.minimal_soz_soni])
        _style(ws)

        ws = wb.create_sheet('Speaking')
        ws.append(['Fan', 'Daraja', 'Mavzu', 'Dars', 'Matn'])
        for item in SpeakingTopshiriq.objects.select_related('dars__mavzu__daraja__fan'):
            lesson = item.dars
            ws.append([lesson.mavzu.daraja.fan.nomi, lesson.mavzu.daraja.nomi, lesson.mavzu.nomi, lesson.sarlavha, item.matn])
        _style(ws)

        stream = BytesIO()
        wb.save(stream)
        log_amal(request.user, 'excel_export', "Ta'lim kontenti Excel formatda yuklandi", request=request)
        response = HttpResponse(stream.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="alaziz_kontent.xlsx"'
        return response


class ContentExcelImportView(APIView):
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload or not upload.name.lower().endswith('.xlsx'):
            return Response({'detail': 'Faqat .xlsx fayl yuboring.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = load_workbook(upload, data_only=True)
        except Exception:
            return Response({'detail': 'Excel faylni o‘qib bo‘lmadi.'}, status=status.HTTP_400_BAD_REQUEST)

        created = {'mavzular': 0, 'testlar': 0, 'listening': 0, 'writing': 0, 'speaking': 0}
        errors = []

        def rows(sheet):
            if sheet not in wb.sheetnames:
                return []
            ws = wb[sheet]
            headers = [str(c.value or '').strip() for c in ws[1]]
            return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(v not in (None, '') for v in row)]

        for index, row in enumerate(rows('Mavzular'), 2):
            try:
                _, _, topic, lesson = _find_course(row['Fan'], row['Daraja'], row['Mavzu'], row['Dars'])
                lesson.tushuntirish_matn = row.get('Tushuntirish') or lesson.tushuntirish_matn
                lesson.misollar = row.get('Misollar') or lesson.misollar
                lesson.tartib = int(row.get('Tartib') or 0)
                lesson.save()
                created['mavzular'] += 1
            except Exception as exc:
                errors.append(f'Mavzular {index}-qator: {exc}')

        for index, row in enumerate(rows('TestSavollar'), 2):
            try:
                _, _, _, lesson = _find_course(row['Fan'], row['Daraja'], row['Mavzu'], row['Dars'])
                exercise, _ = Mashq.objects.get_or_create(dars=lesson, defaults={'sarlavha': f'{lesson.sarlavha} testi', 'otish_bali_foiz': 80})
                question, _ = Savol.objects.get_or_create(mashq=exercise, matn=str(row['Savol']).strip(), defaults={'tartib': int(row.get('Tartib') or 0)})
                correct = str(row.get('Togri') or 'A').strip().upper()
                question.javoblar.all().delete()
                for pos, key in enumerate(['A', 'B', 'C', 'D']):
                    value = row.get(key)
                    if value not in (None, ''):
                        Javob.objects.create(savol=question, matn=str(value), togri=(key == correct), tartib=pos)
                created['testlar'] += 1
            except Exception as exc:
                errors.append(f'TestSavollar {index}-qator: {exc}')

        for index, row in enumerate(rows('Listening'), 2):
            try:
                _, _, _, lesson = _find_course(row['Fan'], row['Daraja'], row['Mavzu'], row['Dars'])
                variants = [str(row.get(k)) for k in ['Variant1', 'Variant2', 'Variant3', 'Variant4'] if row.get(k) not in (None, '')]
                ListeningSavol.objects.update_or_create(dars=lesson, audio_matn=str(row['AudioMatn']).strip(), defaults={'savol': row.get('Savol') or "Eshitganingizni tanlang", 'variantlar': variants, 'togri_javob': row.get('TogriJavob') or '', 'til_kodi': row.get('Til') or 'en-US'})
                created['listening'] += 1
            except Exception as exc:
                errors.append(f'Listening {index}-qator: {exc}')

        for sheet, model, field in [('Writing', WritingTopshiriq, 'Topshiriq'), ('Speaking', SpeakingTopshiriq, 'Matn')]:
            for index, row in enumerate(rows(sheet), 2):
                try:
                    _, _, _, lesson = _find_course(row['Fan'], row['Daraja'], row['Mavzu'], row['Dars'])
                    defaults = {'minimal_soz_soni': int(row.get('MinimalSoz') or 30)} if sheet == 'Writing' else {}
                    model.objects.update_or_create(dars=lesson, matn=str(row[field]).strip(), defaults=defaults)
                    created[sheet.lower()] += 1
                except Exception as exc:
                    errors.append(f'{sheet} {index}-qator: {exc}')

        log_amal(request.user, 'excel_import', str(created), yangi_holat=created, request=request)
        return Response({'import_qilindi': created, 'xatolar': errors[:100]})
