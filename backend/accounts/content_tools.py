import io, json
from django.http import HttpResponse
from django.core import serializers as djserializers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from openpyxl import Workbook, load_workbook
from .permissions import IsAdmin
from courses.models import Fan, Daraja, Mavzu, Dars
from exams.models import Mashq, Savol, Javob, AdminAmalLog

class KontentExcelShablonView(APIView):
    permission_classes=[IsAdmin]
    def get(self, request):
        wb=Workbook(); ws=wb.active; ws.title='Kontent'
        ws.append(['Fan','Daraja','Mavzu','Dars','Savol','A','B','C','D','Togri'])
        ws.append(['English','Beginner','Greetings','Hello','Hello sozi nimani anglatadi?','Salom','Xayr','Rahmat','Iltimos','A'])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        r=HttpResponse(buf.getvalue(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        r['Content-Disposition']='attachment; filename="kontent_shablon.xlsx"'; return r

class KontentExcelImportView(APIView):
    permission_classes=[IsAdmin]; parser_classes=[MultiPartParser]
    def post(self, request):
        file=request.FILES.get('file')
        if not file: return Response({'detail':'Excel fayl tanlang.'},status=400)
        wb=load_workbook(file,data_only=True); ws=wb.active; created=0; errors=[]
        for no,row in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
            fan_n,dar_n,mav_n,dars_n,savol,*rest=row
            if not all([fan_n,dar_n,mav_n,dars_n,savol]): continue
            try:
                fan,_=Fan.objects.get_or_create(nomi=str(fan_n).strip(),defaults={'tartib':Fan.objects.count()+1})
                dar,_=Daraja.objects.get_or_create(fan=fan,nomi=str(dar_n).strip(),defaults={'tartib':fan.darajalar.count()+1})
                mav,_=Mavzu.objects.get_or_create(daraja=dar,nomi=str(mav_n).strip(),defaults={'tartib':dar.mavzular.count()+1})
                dars,_=Dars.objects.get_or_create(mavzu=mav,sarlavha=str(dars_n).strip(),defaults={'tartib':mav.darslar.count()+1,'matn':''})
                mashq,_=Mashq.objects.get_or_create(dars=dars,defaults={'sarlavha':'10 savollik test','otish_bali_foiz':80})
                q,_=Savol.objects.get_or_create(mashq=mashq,matn=str(savol).strip(),defaults={'tartib':mashq.savollar.count()+1})
                variants=rest[:4]; correct=str(rest[4] or 'A').strip().upper()
                for idx,val in enumerate(variants):
                    if val is not None: Javob.objects.get_or_create(savol=q,matn=str(val),defaults={'tartib':idx+1,'togri':correct==chr(65+idx)})
                created+=1
            except Exception as exc: errors.append({'qator':no,'xato':str(exc)[:180]})
        AdminAmalLog.objects.create(foydalanuvchi=request.user,amal='kontent_excel_import',tavsif=f'{created} qator import qilindi')
        return Response({'yaratildi':created,'xatolar':errors})

class BackupJsonView(APIView):
    permission_classes=[IsAdmin]
    def get(self, request):
        from django.apps import apps
        models=[]
        for app_label in ('accounts','courses','exams'):
            models.extend(apps.get_app_config(app_label).get_models())
        data=djserializers.serialize('json',[obj for model in models for obj in model.objects.all()],indent=2)
        AdminAmalLog.objects.create(foydalanuvchi=request.user,amal='backup_yuklandi',tavsif='JSON backup yuklandi')
        r=HttpResponse(data,content_type='application/json'); r['Content-Disposition']='attachment; filename="alaziz_backup.json"'; return r
